import pandas as pd
import re
import openpyxl
import csv
import os
from openpyxl.utils import get_column_letter

file_name = 'sample-input/inefficient_vms_report.csv'

save_file_name = 'sample-output/inefficient_vms_report_beautified.xlsx'

df = pd.read_csv(file_name)

unique_clusters = df['Cluster'].unique()

clusters = {}

for cluster in unique_clusters :
    clusters[cluster] = []

for i in range(len(df)) :
    current_cluster = df.loc[i, 'Cluster']
    for cluster in unique_clusters :
        if current_cluster == cluster :
            clusters[cluster].append(df.iloc[i])

for cluster, cluster_info in clusters.items() :
    cluster = cluster + '.csv'
    if not os.path.exists(cluster) :
        with open(cluster, 'w') as file :
            writer = csv.writer(file)
            header_row = ['Name', 'Efficiency', 'Efficiency Detail', 'Project', 'Owner', 'Cluster']
            writer.writerow(header_row)
            for row in cluster_info :
                writer.writerow(row)

table_name_counter = 0

def spacer() :
    for i in range(5) :
        sheet.append([])

def yellow_text(current_header_row) :
    for cell in sheet[current_header_row] :
        cell.font = openpyxl.styles.Font(bold = True)
        cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def bold_text(current_header_row) :
    for cell in sheet[current_header_row] :
        cell.font = openpyxl.styles.Font(bold = True)

def make_table(start_row, end_row, num_cols) :

    global table_name_counter

    table_name = f"Table{table_name_counter}"

    if num_cols == 4:
        table_range = f'A{start_row}:D{end_row}'
    elif num_cols == 3 :
        table_range = f'A{start_row}:C{end_row}'
    
    table = openpyxl.worksheet.table.Table(displayName=table_name, ref=table_range)

    style = openpyxl.worksheet.table.TableStyleInfo(
        name="TableStyleMedium4", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )

    table.tableStyleInfo = style

    for col in range(num_cols + 1, 5):
        cell = sheet[f"{get_column_letter(col)}{start_row}"]
        cell.fill = openpyxl.styles.PatternFill(fill_type=None)
        cell.border = None
        cell.font = openpyxl.styles.Font(color="000000")

    sheet.add_table(table)

    table_name_counter = table_name_counter + 1

workbook = openpyxl.Workbook()

try:
    workbook = openpyxl.load_workbook(save_file_name)

except FileNotFoundError:
    workbook.save(save_file_name)

df = pd.DataFrame()

sheet = workbook.active

for cluster in unique_clusters :
    cluster = cluster + '.csv'
    if (os.path.exists(cluster)) :
        df = pd.read_csv(cluster)
        cluster = cluster[:-4]
        sheet = workbook.create_sheet(cluster)

        overprovisioned_vms = []
        inactive_vms = []
        constrained_overprovisioned_vms = []
        constrained_vms = []

        for i in range(len(df)) :
            
            vm_name = df.loc[i, 'Name']
            efficiency = df.loc[i, 'Efficiency']
            sentence = df.loc[i, 'Efficiency Detail']
            new_row = []

            if(efficiency == 'Overprovisioned') :
                cpu_match = re.search(r'hourly average CPU usage \((\d+(?:\.\d+)?)%\).*?\b99\.5% of the time', sentence)
                memory_match = re.search(r'hourly average memory usage \((\d+(?:\.\d+)?)%\).*?\b99\.5% of the time', sentence)

                if cpu_match and memory_match:
                    cpu_value = float(cpu_match.group(1))
                    memory_value = float(memory_match.group(1))
                    new_row = [vm_name, efficiency, cpu_value, memory_value]
                    overprovisioned_vms.append(new_row)

                elif cpu_match:
                    cpu_value = float(cpu_match.group(1))
                    new_row = [vm_name, efficiency, cpu_value, None]
                    overprovisioned_vms.append(new_row)

                elif memory_match:
                    memory_value = float(memory_match.group(1))
                    new_row = [vm_name, efficiency, None, memory_value]
                    overprovisioned_vms.append(new_row)
                
            elif (efficiency == 'Inactive') :
                inactive_match = re.search(r'Inactive: (.*)', sentence)

                if inactive_match:
                    inactive_value = inactive_match.group(1)
                    new_row = [vm_name, efficiency, inactive_value, None]
                    inactive_vms.append(new_row)
                        
            elif (efficiency == 'Overprovisioned,Constrained') :
                cpu_match_overprovisioned = re.search(r'hourly average CPU usage \((\d+(?:\.\d+)?)%\).*?\b99\.5% of the time', sentence)
                memory_match_overprovisioned = re.search(r'hourly average memory usage \((\d+(?:\.\d+)?)%\).*?\b99\.5% of the time', sentence)
                cpu_match_constrained = re.search(r'hourly average CPU usage \((\d+(?:\.\d+)?)%\).*?\b5% of the time', sentence)
                memory_match_constrained = re.search(r'hourly average memory usage \((\d+(?:\.\d+)?)%\).*?\b5% of the time', sentence)

                if cpu_match_overprovisioned and memory_match_overprovisioned :
                    cpu_match_overprovisioned_value = float(cpu_match_overprovisioned.group(1))
                    memory_match_overprovisioned_value = float(memory_match_overprovisioned.group(1))
                    new_row = [vm_name, efficiency, cpu_match_overprovisioned_value, memory_match_overprovisioned_value]
                    constrained_overprovisioned_vms.append(new_row)
                
                elif cpu_match_overprovisioned and memory_match_constrained :
                    cpu_match_overprovisioned_value = float(cpu_match_overprovisioned.group(1))
                    memory_match_constrained_value = float(memory_match_constrained.group(1))
                    new_row = [vm_name, efficiency, cpu_match_overprovisioned_value, memory_match_constrained_value]
                    constrained_overprovisioned_vms.append(new_row)

                elif cpu_match_constrained and memory_match_overprovisioned :
                    cpu_match_constrained_value = float(cpu_match_constrained.group(1))
                    memory_match_overprovisioned_value = float(memory_match_overprovisioned.group(1))
                    new_row = [vm_name, efficiency, cpu_match_constrained_value, memory_match_overprovisioned_value]
                    constrained_overprovisioned_vms.append(new_row)

                elif cpu_match_constrained and memory_match_constrained :
                    cpu_match_constrained_value = float(cpu_match_constrained.group(1))
                    memory_match_constrained_value = float(memory_match_constrained.group(1))
                    new_row = [vm_name, efficiency, cpu_match_constrained_value, memory_match_constrained_value]
                    constrained_overprovisioned_vms.append(new_row)

            elif (efficiency == 'Constrained') :
                cpu_match_constrained = re.search(r'hourly average CPU usage \((\d+(?:\.\d+)?)%\).*?\b5% of the time', sentence)
                memory_match_constrained = re.search(r'hourly average memory usage \((\d+(?:\.\d+)?)%\).*?\b5% of the time', sentence)

                if cpu_match_constrained and memory_match_constrained :
                    cpu_match_constrained_value = float(cpu_match_constrained.group(1))
                    memory_match_constrained_value = float(memory_match_constrained.group(1))
                    new_row = [vm_name, efficiency, cpu_match_constrained_value, memory_match_constrained_value]
                    constrained_vms.append(new_row)

                elif cpu_match_constrained :
                    cpu_match_constrained_value = float(cpu_match_constrained.group(1))
                    new_row = [vm_name, efficiency, cpu_match_constrained_value, None]
                    constrained_vms.append(new_row)

                elif memory_match_constrained :
                    memory_match_constrained_value = float(memory_match_constrained.group(1))
                    new_row = [vm_name, efficiency, None, memory_match_constrained_value]
                    constrained_vms.append(new_row)

        if(overprovisioned_vms) :
            overprovisioned_vms_heading = f'***** Overprovisioned VMs (count: {len(overprovisioned_vms)}) *****'
            overprovisioned_vms_desc1 = 'For Overprovisioned VMs, the % values listed under the respective CPU/Memory usage columns are the maximum hourly average values for 99.5% of the time in the lookback period'
            overprovisioned_vms_desc2 = 'Blank cell values indicate that the VM is not inefficient for that particular metric (CPU/Memory)'
            sheet.append([overprovisioned_vms_heading])
            bold_text(sheet.max_row)
            sheet.append([overprovisioned_vms_desc1])
            bold_text(sheet.max_row)
            sheet.append([overprovisioned_vms_desc2])
            bold_text(sheet.max_row)
            sheet.append([])
            overprovisioned_vms_header = ['VM Name', 'Efficiency', 'CPU Usage % (Max Hourly Average)', 'Memory Usage % (Max Hourly Average)']
            sheet.append(overprovisioned_vms_header)
            current_header_row = sheet.max_row
            yellow_text(current_header_row)

            for row in overprovisioned_vms :
                sheet.append(row)

            current_footer_row = sheet.max_row
            if (overprovisioned_vms) :
                make_table(current_header_row, current_footer_row, 4)

            spacer()
        
        else :
            sheet.append(['***** No Overprovisioned VMs in this cluster *****'])
            bold_text(sheet.max_row)

            spacer()

        if(constrained_overprovisioned_vms) :
            constrained_overprovisioned_vms_heading = f'***** Overprovisioned and/or Constrained VMs (count: {len(constrained_overprovisioned_vms)}) *****'
            constrained_overprovisioned_vms_desc1 = 'For Overprovisioned VMs, the % values listed under the respective CPU/Memory usage columns are the maximum hourly average values for 99.5% of the time in the lookback period'
            constrained_overprovisioned_vms_desc2 = 'For Constrained VMs, the % values listed under the respective CPU/Memory usage columns are the maximum hourly average values for 5% of the time in the lookback period'
            constrained_overprovisioned_vms_desc3 = 'Blank cell values indicate that the VM is not inefficient for that particular metric (CPU/Memory)'
            sheet.append([constrained_overprovisioned_vms_heading])
            bold_text(sheet.max_row)
            sheet.append([constrained_overprovisioned_vms_desc1])
            bold_text(sheet.max_row)
            sheet.append([constrained_overprovisioned_vms_desc2])
            bold_text(sheet.max_row)
            sheet.append([constrained_overprovisioned_vms_desc3])
            bold_text(sheet.max_row)
            sheet.append([])
            constrained_overprovisioned_vms_header = ['VM Name', 'Efficiency', 'CPU Usage % (Max Hourly Average)', 'Memory Usage % (Max Hourly Average)']
            sheet.append(constrained_overprovisioned_vms_header)
            current_header_row = sheet.max_row
            yellow_text(current_header_row)

            for row in constrained_overprovisioned_vms :
                sheet.append(row)

            current_footer_row = sheet.max_row
            if (constrained_overprovisioned_vms) :
                make_table(current_header_row, current_footer_row, 4)

            spacer()
        
        else :
            sheet.append(['***** No Overprovisioned and/or Constrained VMs in this cluster *****'])
            bold_text(sheet.max_row)

            spacer()

        if (constrained_vms) :
            constrained_vms_heading = f'***** Constrained VMs (count: {len(constrained_vms)}) *****'
            constrained_vms_desc1 = ' For Constrained VMs, the % values listed under the respective CPU/Memory usage columns are the maximum hourly average values for 5% of the time in the lookback period'
            constrained_vms_desc2 = 'Blank cell values indicate that the VM is not inefficient for that particular metric (CPU/Memory)'
            sheet.append([constrained_vms_heading])
            bold_text(sheet.max_row)
            sheet.append([constrained_vms_desc1])
            bold_text(sheet.max_row)
            sheet.append([constrained_vms_desc2])
            bold_text(sheet.max_row)
            sheet.append([])
            constrained_vms_header = ['VM Name', 'Efficiency', 'CPU Usage % (Max Hourly Average)', 'Memory Usage % (Max Hourly Average)']
            sheet.append(constrained_vms_header)
            current_header_row = sheet.max_row
            yellow_text(current_header_row)

            for row in constrained_vms :
                sheet.append(row)

            current_footer_row = sheet.max_row

            make_table(current_header_row, current_footer_row, 4)

            spacer()
        else :
            sheet.append(['***** No Constrained VMs in this cluster *****'])
            bold_text(sheet.max_row)

            spacer()
        
        if(inactive_vms) :
            inactive_vms_heading = f'***** Inactive VMs (count: {len(inactive_vms)}) *****'
            sheet.append([inactive_vms_heading])
            bold_text(sheet.max_row)
            sheet.append([])
            inactive_vms_header = ['VM Name', 'Efficiency', 'Efficiency Detail']
            sheet.append(inactive_vms_header)
            current_header_row = sheet.max_row
            yellow_text(current_header_row)

            for row in inactive_vms :
                sheet.append(row)

            current_footer_row = sheet.max_row
            if (inactive_vms) :
                make_table(current_header_row, current_footer_row, 3)
        
        else :
            sheet.append(['***** No Inactive VMs in this cluster *****'])
            bold_text(sheet.max_row)

            spacer()

        column_width = 35

        for col in sheet.columns:
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = column_width

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='left')

        workbook.save(save_file_name)
    
for cluster in unique_clusters :
    cluster = cluster + '.csv'
    if (os.path.exists(cluster)) :
        os.remove(cluster)

if('Sheet' in workbook.sheetnames) :
    sheet = workbook['Sheet']
    workbook.remove(sheet)
    workbook.save(save_file_name)